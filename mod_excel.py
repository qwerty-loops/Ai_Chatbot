import pandas as pd
import openpyxl
import os
import warnings
warnings.filterwarnings("ignore")

def process_excel(excel_link, output_path):
    print(f"Processing Excel file: {excel_link}")
    try:
        # Check if the file exists
        if not os.path.exists(excel_link):
            raise FileNotFoundError(f"The file {excel_link} does not exist.")
        
        # Load the workbook twice: once with data_only=True and once without
        workbook = openpyxl.load_workbook(excel_link, data_only=True)
        workbook_with_formulas = openpyxl.load_workbook(excel_link, data_only=False)

        # Define the target RGB color for yellow
        target_rgb = 'FFFFFF00'

        # DataFrame to store yellow cells information
        yellow_data = []

        # Iterate through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_with_formulas = workbook_with_formulas[sheet_name]

            # Retrieve cells with the specific yellow fill
            for row in sheet.iter_rows():
                for cell in row:
                    fill = cell.fill
                    fg_color = fill.fgColor
                    if fg_color.type == 'rgb' and fg_color.rgb == target_rgb:
                        left_cell_value = sheet.cell(row=cell.row, column=cell.column - 2).value if cell.column > 1 else None
                        formula = sheet_with_formulas[cell.coordinate].value
                        yellow_data.append({
                            "Sheet": sheet_name,
                            "Cell": cell.coordinate,
                            "Left Cell Value": left_cell_value,
                            "Value": cell.value,
                            "Formula": formula
                        })

        # Create a DataFrame from the yellow cells information
        yellow_df = pd.DataFrame(yellow_data)

        # Add the results sheet to the original workbook
        if 'Results' in workbook.sheetnames:
            del workbook['Results']  # Delete the existing results sheet if it already exists
        results_sheet = workbook.create_sheet(title='Results')
        results_sheet.append(list(yellow_df.columns))  # Add headers
        for row in yellow_df.itertuples(index=False, name=None):
            results_sheet.append(row)

        # Save the updated workbook as output.xlsx
        workbook.save(output_path)

        return f"Processed Excel file saved as {output_path}"

    except Exception as e:
        print(e)
        return str(e)

# Prompt the user for the input Excel file path and the output file path
excel_link = input("Enter the file path to the desired Excel : ").strip('\'"')
output_path = input("Enter the file path to save the output Excel : ").strip('\'"')

# Execute the function with user-provided paths
print(process_excel(excel_link, output_path))
