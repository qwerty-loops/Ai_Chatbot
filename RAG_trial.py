import openai
import os
import pandas as pd
import warnings
from dotenv import load_dotenv
from langchain_community.chat_models import ChatOpenAI
from langchain_community.embeddings import OpenAIEmbeddings
from langchain.prompts import PromptTemplate
from langchain.memory import ConversationBufferWindowMemory
from langchain_community.vectorstores import FAISS
from langchain.schema import HumanMessage
import openpyxl

# Ignore specific warnings
warnings.filterwarnings("ignore", category=DeprecationWarning, module="langchain")

# Load environment variables from .env file
load_dotenv()

# Set up your OpenAI API key
openai.api_key = os.getenv("OPENAI_API_KEY")

# Function to load and convert Excel to text
def load_excel(file_path):
    workbook = pd.ExcelFile(file_path)
    text = ""
    for sheet_name in workbook.sheet_names:
        try:
            sheet = pd.read_excel(workbook, sheet_name=sheet_name)
            text += f"Sheet name: {sheet_name}\n"
            text += sheet.to_string(index=False, header=True)
            text += "\n\n"
        except Exception as e:
            text += f"Error reading {sheet_name}: {e}\n\n"
    return text

# Function to load document based on file extension
def load_document(file_path):
    if file_path.endswith('.xlsx'):
        return load_excel(file_path)
    else:
        raise ValueError("Unsupported file type. Please provide an Excel file.")

# Function to process Excel and generate output
def process_excel(excel_link, output_path):
    print(f"Processing Excel file: {excel_link}")
    try:
        # Check if the file exists
        if not os.path.exists(excel_link):
            raise FileNotFoundError(f"The file {excel_link} does not exist.")
        
        # Load the workbook twice: once with data_only=True and once without
        workbook = openpyxl.load_workbook(excel_link, data_only=True)
        workbook_with_formulas = openpyxl.load_workbook(excel_link, data_only=False)

        # Define the target RGB color for yellow
        target_rgb = 'FFFFFF00'

        # DataFrame to store yellow cells information
        yellow_data = []

        # Iterate through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_with_formulas = workbook_with_formulas[sheet_name]

            # Retrieve cells with the specific yellow fill
            for row in sheet.iter_rows():
                for cell in row:
                    fill = cell.fill
                    fg_color = fill.fgColor
                    if fg_color.type == 'rgb' and fg_color.rgb == target_rgb:
                        left_cell_value = sheet.cell(row=cell.row, column=cell.column - 2).value if cell.column > 1 else None
                        formula = sheet_with_formulas[cell.coordinate].value
                        yellow_data.append({
                            "Sheet": sheet_name,
                            "Cell": cell.coordinate,
                            "Left Cell Value": left_cell_value,
                            "Value": cell.value,
                            "Formula": formula
                        })

        # Create a DataFrame from the yellow cells information
        yellow_df = pd.DataFrame(yellow_data)

        # Add the results sheet to the original workbook
        if 'Results' in workbook.sheetnames:
            del workbook['Results']  # Delete the existing results sheet if it already exists
        results_sheet = workbook.create_sheet(title='Results')
        results_sheet.append(list(yellow_df.columns))  # Add headers
        for row in yellow_df.itertuples(index=False, name=None):
            results_sheet.append(row)

        # Save the updated workbook as output.xlsx
        workbook.save(output_path)

        return f"Processed Excel file saved as {output_path}"

    except Exception as e:
        print(e)
        return str(e)

# Function to create embeddings and vector store for document retrieval
def create_vector_store(file_path):
    document_text = load_document(file_path)
    embeddings = OpenAIEmbeddings(openai_api_key=openai.api_key)
    return FAISS.from_texts([document_text], embeddings)

# Prompt the user to enter the file path
print("Welcome to NoteSight AI Assistant!")

while True:
    try:
        file_path = input("Please enter the file path: ").strip('\'"')
        vector_store = create_vector_store(file_path)
        break
    except ValueError as e:
        print(e)

# Define a prompt template
prompt_template = """
You are a helpful assistant. 
If you have any questions, feel free to ask.
If you are unsure of the answer, you can say "I'm sorry I didn't understand the question".
Answer the following question based on the provided context, which shall be an uploaded excel document.:

{chat_history}
{context}
Question: {question}

Answer:
"""

# Create a LangChain prompt template
template = PromptTemplate(template=prompt_template, input_variables=["question", "chat_history", "context"])

# Create a ChatOpenAI instance
llm = ChatOpenAI(model_name="gpt-4-turbo", temperature=0, verbose=True)

# Initialize conversational memory
conversational_memory = ConversationBufferWindowMemory(
    memory_key='chat_history',
    k=7,
    return_messages=True
)

# Create a class for the AI agent and define a run method
class AIAgent:
    def __init__(self, template, llm, memory, retriever):
        self.template = template
        self.llm = llm
        self.memory = memory
        self.retriever = retriever

    def run(self, prompt):
        # Retrieve chat history
        chat_history = self.memory.load_memory_variables({})

        # Retrieve relevant documents
        docs = self.retriever.get_relevant_documents(prompt)
        context = "\n".join([doc.page_content for doc in docs])

        # Format the prompt with chat history and context
        formatted_prompt = self.template.format(question=prompt, chat_history=chat_history.get('chat_history', ''), context=context)
        messages = [HumanMessage(content=formatted_prompt)]
        
        response = self.llm(messages)

        # Update memory with the latest interaction
        self.memory.save_context({"question": prompt}, {"response": response.content})

        return response.content

# Initialize the retriever
retriever = vector_store.as_retriever()

# Usage
agent = AIAgent(template, llm, conversational_memory, retriever)

while True:
    option = input("Enter 'ask' to ask a question, 'generate' to generate output Excel, or 'exit' to quit: ").strip().lower()
    
    if option == "exit":
        print("Goodbye!")
        break
    elif option == "ask":
        question = input("Ask your question: ")
        response = agent.run(question)
        print(f"Assistant: {response}")
    elif option == "generate":
        output_path = input("Enter the file path to save the output Excel: ").strip('\'"')
        result = process_excel(file_path, output_path)
        print(result)
        
        # Update the vector store with the new file
        try:
            vector_store = create_vector_store(output_path)
            agent.retriever = vector_store.as_retriever()
            file_path = output_path  # Update the file path to the new output file
        except ValueError as e:
            print(e)
    else:
        print("Invalid option. Please try again.")
